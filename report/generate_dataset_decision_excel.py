#!/usr/bin/env python3
"""生成《机器人数据链路与重要数据集深度调研报告》配套 Excel 决策速查表。

Tab 顺序（老板阅读路径）：
  自建路线图 → 训练范式 → 数据集目录 → 链路与工具 → 行业对标（人形）

排序规则（Excel 独立编排，不强行对齐 HTML 报告）：
  · 3-数据集目录：真机 traj → 仿真 traj → 合成 traj → 历史/不推荐；组内按年份、同 year 按规模
  · 4-链路与工具：A 建链路（按 Phase）→ B 消费端（IL → VLA 桌面 → VLA 人形）
  · 2-训练范式：数据门槛从低到高，首列「推荐阶段」与路线图 Phase 闭环
  · 5-行业对标：中国·可对标 → 中国·知存在 → 海外·头部 → 海外/车企；主表瘦身+中文名检索
  · 各表含「落地推荐/落地借鉴」(★★/★/—/不推荐) 与「选择理由」；★★ 浅蓝加粗，不推荐整行灰显
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "机器人数据链路数据集决策速查表.xlsx"

# Tab 名称（交叉引用统一用 Tab 名，避免调序后错乱）
T_ROAD = "1-自建路线图"
T_PARADIGM = "2-训练范式速查"
T_DATASET = "3-数据集目录"
T_PIPE = "4-链路与工具"
T_INDUSTRY = "5-行业对标（人形）"

# 配色（与 HTML 报告主色接近）
C_HEADER = "1F3F68"
C_HEADER_FONT = "FFFFFF"
C_ROW_A = "FFFFFF"
C_ROW_B = "F4F7FB"
C_PRIO_P0 = "EAF2FC"
C_PRIO_P1 = "EAF7EF"
C_PRIO_P3 = "FFF8EE"
C_ACCENT = "0C8796"
C_MUTED = "A0A8B4"
C_MUTED_BG = "ECEFF3"
C_PRIMARY_BG = "EAF2FC"
C_SECONDARY_BG = "F4F7FB"
C_BANNER_BG = "E8F4F4"
C_BANNER_FONT = "1F3F68"
C_HDR_BUILD = "2A5080"  # 在建链路表头
C_HDR_ROUTE = "1F5FAE"  # 技术路线表头
C_HDR_REUSE = "138A5B"  # 可复用
C_HDR_LEARN = "0C8796"  # 可借鉴
C_HDR_AVOID = "8B3A3A"  # 勿照搬
C_HDR_ACTION = "0A6B75"  # 建议动作
C_CELL_BUILD = "D6E6F5"
C_CELL_ROUTE = "E3EEF9"
C_CELL_REUSE = "DFF3E8"
C_CELL_LEARN = "D9F0F0"
C_CELL_AVOID = "FCE8E8"
C_CELL_ACTION = "DFF5F5"

THIN = Side(style="thin", color="D9E2EE")
MEDIUM = Side(style="medium", color="1F3F68")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SECTION_BORDER = Border(left=THIN, right=THIN, top=MEDIUM, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor=C_HEADER)
HEADER_FONT = Font(name="PingFang SC", bold=True, color=C_HEADER_FONT, size=10)
BODY_FONT = Font(name="PingFang SC", size=10)
MUTED_FONT = Font(name="PingFang SC", size=10, color=C_MUTED)
NAME_PRIMARY_FONT = Font(name="PingFang SC", size=11, bold=True, color=C_HEADER)
NAME_SECONDARY_FONT = Font(name="PingFang SC", size=10, bold=True, color=C_HEADER)
RECOMMEND_FONT = Font(name="PingFang SC", size=10, bold=True, color=C_ACCENT)
BANNER_FONT = Font(name="PingFang SC", size=9, color=C_BANNER_FONT)
FONT_BUILD = Font(name="PingFang SC", size=10, bold=True, color=C_HEADER)
FONT_ROUTE = Font(name="PingFang SC", size=10, bold=True, color="1F5FAE")
FONT_REUSE = Font(name="PingFang SC", size=10, bold=True, color=C_HDR_REUSE)
FONT_LEARN = Font(name="PingFang SC", size=10, bold=True, color=C_HDR_LEARN)
FONT_AVOID = Font(name="PingFang SC", size=10, bold=True, color=C_HDR_AVOID)
FONT_ACTION = Font(name="PingFang SC", size=10, bold=True, color=C_HDR_ACTION)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

# 各 Sheet 列宽（字符）
COL_WIDTHS = {
    T_ROAD: [8, 10, 30, 22, 28, 26, 40, 10, 10, 36],
    T_PARADIGM: [16, 8, 34, 22, 24, 32, 28, 22, 22, 28],
    T_DATASET: [8, 34, 18, 22, 11, 14, 20, 22, 14, 22, 8, 18, 22, 8, 8, 18, 14, 24, 22, 20, 22, 22],
    T_PIPE: [8, 34, 12, 22, 16, 22, 28, 28, 18, 12, 24, 24, 26],
    T_INDUSTRY: [8, 22, 12, 22, 32, 24, 11, 16, 24, 24, 22, 26, 12],
}

# Sheet5 重点列视觉：表头色 / 单元格底色 / 字重色
INDUSTRY_HEADER_FILLS = {
    5: C_HDR_BUILD,
    6: C_HDR_ROUTE,
    9: C_HDR_REUSE,
    10: C_HDR_LEARN,
    11: C_HDR_AVOID,
    12: C_HDR_ACTION,
}
INDUSTRY_EMPHASIS = {
    5: (C_CELL_BUILD, FONT_BUILD),
    6: (C_CELL_ROUTE, FONT_ROUTE),
    9: (C_CELL_REUSE, FONT_REUSE),
    10: (C_CELL_LEARN, FONT_LEARN),
    11: (C_CELL_AVOID, FONT_AVOID),
    12: (C_CELL_ACTION, FONT_ACTION),
}


def _tier_from_recommend(value: object) -> str:
    s = str(value or "").strip()
    if not s or s in {"—", "-"}:
        return "normal"
    if "不推荐" in s:
        return "muted"
    if s.startswith("★★"):
        return "primary"
    if s.startswith("★"):
        return "secondary"
    return "normal"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _write_merged_banner(ws, row: int, ncols: int, text: str, *, height: int = 36) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = _fill(C_BANNER_BG)
    cell.font = BANNER_FONT
    cell.alignment = WRAP
    cell.border = BORDER
    for col_idx in range(2, ncols + 1):
        c = ws.cell(row=row, column=col_idx)
        c.fill = _fill(C_BANNER_BG)
        c.border = BORDER
    ws.row_dimensions[row].height = height


def style_table(
    ws,
    headers: list[str],
    rows: list[list],
    *,
    col_widths: list[int] | None = None,
    freeze: str = "B2",
    priority_col: int | None = None,
    name_col: int = 1,
    recommend_col: int | None = None,
    category_col: int | None = None,
    section_starts: list[int] | None = None,
    banner: str | None = None,
    banner2: str | None = None,
    banner_height: int = 48,
    banner2_height: int = 40,
    data_row_height: int = 52,
    auto_filter: bool = False,
    header_fills: dict[int, str] | None = None,
    emphasis_cols: dict[int, tuple[str, Font]] | None = None,
) -> None:
    ncols = len(headers)
    header_row = 1
    data_start = 2
    header_fills = header_fills or {}
    emphasis_cols = emphasis_cols or {}

    if banner:
        _write_merged_banner(ws, 1, ncols, banner, height=banner_height)
        header_row = 2
        data_start = 3
        if banner2:
            _write_merged_banner(ws, 2, ncols, banner2, height=banner2_height)
            header_row = 3
            data_start = 4

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.fill = _fill(header_fills.get(col_idx, C_HEADER))
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 32

    # section_starts: 1-based index into `rows`（第 1 条数据 = 1）
    section_excel_rows = {data_start - 1 + s for s in (section_starts or [])}

    for offset, row in enumerate(rows):
        row_idx = data_start + offset
        tier = (
            _tier_from_recommend(row[recommend_col - 1])
            if recommend_col is not None and recommend_col <= len(row)
            else "normal"
        )
        category = (
            str(row[category_col - 1] or "")
            if category_col is not None and category_col <= len(row)
            else ""
        )

        bg = C_ROW_A if row_idx % 2 == 0 else C_ROW_B
        if tier == "primary":
            bg = C_PRIMARY_BG
        elif tier == "muted":
            bg = C_MUTED_BG
        elif category.startswith("中国·可对标") and tier == "secondary":
            bg = C_PRIO_P1
        elif category.startswith("中国·知存在"):
            bg = C_SECONDARY_BG
        elif category.startswith("车企"):
            bg = C_PRIO_P3
        elif priority_col is not None and priority_col <= len(row):
            prio = str(row[priority_col - 1] or "")
            if prio == "P0":
                bg = C_PRIO_P0
            elif prio == "P1":
                bg = C_PRIO_P1
            elif prio.startswith("P3") or prio == "P2":
                bg = C_PRIO_P3

        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row[col_idx - 1] if col_idx - 1 < len(row) else None
            cell.value = value
            cell.alignment = CENTER if col_idx in {recommend_col, category_col} else WRAP
            cell.border = SECTION_BORDER if row_idx in section_excel_rows else BORDER

            emp = emphasis_cols.get(col_idx) if tier != "muted" else None
            if emp:
                cell.fill = _fill(emp[0])
            else:
                cell.fill = _fill(bg)

            if tier == "muted":
                cell.font = MUTED_FONT
            elif emp:
                cell.font = emp[1]
            elif col_idx == recommend_col:
                cell.font = RECOMMEND_FONT if tier in {"primary", "secondary"} else BODY_FONT
            elif col_idx == name_col:
                if tier == "primary":
                    cell.font = NAME_PRIMARY_FONT
                elif tier == "secondary":
                    cell.font = NAME_SECONDARY_FONT
                else:
                    cell.font = BODY_FONT
            else:
                cell.font = BODY_FONT

        ws.row_dimensions[row_idx].height = data_row_height

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col_idx in range(1, ncols + 1):
            letter = get_column_letter(col_idx)
            max_len = len(str(headers[col_idx - 1]))
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, min(len(str(row[col_idx - 1] or "")), 50))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 44)

    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    if auto_filter:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(ncols)}{data_start + len(rows) - 1}"
        )


# ---------------------------------------------------------------------------
# 3-数据集目录
# ---------------------------------------------------------------------------
DATASET_HEADERS = [
    "落地推荐",
    "选择理由",
    "数据集名称",
    "机构/团队",
    "发布年份",
    "开源/可得性",
    "规模",
    "数据内容类型",
    "是否含 robot action",
    "主要视角",
    "真机/仿真",
    "机器人类型",
    "模态",
    "语言标注",
    "失败样本",
    "Action 类型 / 坐标系",
    "数据格式",
    "适用任务/技能",
    "适用训练范式",
    "常用消费模型/框架",
    "主要局限",
    "参考价值",
]

# 排序：真机 traj → 仿真 traj → 合成 traj → 历史/不推荐；组内按年份、同 year 按规模
DATASET_ROWS = [
    # ── 真机 traj ──
    [
        "★★",
        "低成本单臂 + RLDS 导出成熟；与 Octo/OpenVLA/RT-1 生态绑定最深，小团队 Phase 2 首选公开对标",
        "BridgeData V2",
        "UC Berkeley / RAIL",
        "2023 (CoRL)",
        "开源",
        "60,096 traj / 24 envs / 13 skills",
        "真机机器人遥操作轨迹",
        "是",
        "第三人称 fixed + 腕部 camera",
        "真机",
        "WidowX 250 单臂",
        "RGB、wrist RGB、proprio、EEF delta、language、success",
        "有",
        "少",
        "EEF delta pose + gripper / end-effector (6+1)",
        "RLDS / TFDS",
        "桌面 pick/place/push/stack/open drawer；低成本单站采集",
        "VLA预训练(经OXE)；VLA微调；单臂 IL",
        "Octo / OpenVLA / RT-1",
        "单臂桌面；与 Franka/人形 action 空间不同",
        "早期团队 teleop + RLDS 导出首选对标",
    ],
    [
        "—",
        "多模态/力触觉研究向；标定同步极复杂，非标准 VLA 开箱路径，默认落地可跳过",
        "RH20T",
        "清华大学等",
        "2023 (CoRL/RSS)",
        "开源",
        "110k+ seq / ~20TB / 50M+ frames / ~150 skills",
        "真机机器人遥操作轨迹；人类第三人称示范视频（部分序列，无 robot action）",
        "是（主序列）；人类 demo 视频无 robot action",
        "8–10 路 global RGBD + in-hand；人类 demo 为第三人称 exo（非 ego）",
        "真机",
        "多种机械臂/夹爪/传感器配置",
        "RGBD、IR、FT、tactile、audio、joint/EEF/gripper action、language",
        "有",
        "少",
        "joint + EEF + gripper + FT / mixed 多维",
        "自定义多模态",
        "接触丰富 manipulation；one-shot imitation；力/触觉/多模态 IL",
        "多模态 IL 研究；传感器融合研究（非标准 VLA 开箱即用）",
        "多模态 IL 研究框架",
        "标定与同步极复杂；数据量大；schema 统一难",
        "接触/力控任务的多模态同步链路参考",
    ],
    [
        "★★",
        "分布式真机采集协议与 metadata 设计标杆；Franka / 多场景泛化团队首选对标",
        "DROID",
        "Stanford / Berkeley / TRI / Google DeepMind 等",
        "2024 (RSS)",
        "开源",
        "76k traj / ~350h / 564 scenes / 86 tasks",
        "真机机器人遥操作轨迹",
        "是",
        "第三人称 fixed ×3 + depth（无 wrist 为主视角）",
        "真机",
        "Franka Panda（统一硬件栈）",
        "RGB×3、depth、calib、joint action、language",
        "有",
        "少",
        "joint position + gripper / joint space (7+1)",
        "自定义 + camera calib metadata",
        "in-the-wild manipulation；pick/place/push；真实场景泛化",
        "IL微调(需adapter)；VLA微调(需adapter)；场景泛化研究",
        "自定义 IL / VLA adapter",
        "单 embodiment；分布式标注一致性；长尾任务采样不均",
        "多地点采集协议与 metadata 标杆",
    ],
    [
        "★",
        "国内可复用人形公开集+taxonomy/QA 对标；硬件绑定强，通用桌面团队慎作主训。行业路径见「5-行业对标」智元 ★★",
        "AgiBot World",
        "AgiBot / OpenDriveLab 等",
        "2025",
        "部分公开（Alpha/Beta/Colosseo 需核实版本）",
        "~1M traj / ~2976h / 217 tasks / 3000+ objects",
        "真机机器人遥操作轨迹",
        "是",
        "RGBD + visuo-tactile（第三人称为主）",
        "真机",
        "AgiBot G1 移动双臂人形 + 灵巧手",
        "RGBD、visuo-tactile、dexterous action、language、object/scene meta",
        "有",
        "未明确统一",
        "dexterous hand + arm joints / joint & EEF 高维",
        "AgiBot 自定义（多版本）",
        "长时序任务；工具使用；双臂协作；人形家居/工业场景",
        "人形/灵巧手 VLA 预训练与微调",
        "GO-1 / ViLLA",
        "版本口径不一；完整可得性需核实；硬件同质集群",
        f"可复用公开子集作 taxonomy/QA；完整路径见 {T_INDUSTRY}: 智元 AgiBot",
    ],
    [
        "★",
        "国内 failure + multi-embodiment 统一协议参考；baseline 稳定后 Phase 5a 扩展，非 Phase 2 起点",
        "RoboMIND（真机 + failure）",
        "X-Humanoid / RoboMIND 团队",
        "2025 (RSS)",
        "开源",
        "107k traj / 479 tasks / 5k failure demos / 4 embodiments",
        "真机机器人遥操作轨迹；含 failure demonstration",
        "是",
        "teleop 采集（第三人称为主，依 embodiment 配置）",
        "真机（另含 Isaac Sim 轨迹，见同表仿真组）",
        "Franka、UR5e、Cobot Magic 双臂、Tien Kung 人形",
        "teleop traj、language、failure label、object/skill taxonomy",
        "有",
        "有（约 5k failure）",
        "per-embodiment teleop action / 各异（统一协议下采集）",
        "RoboMIND 统一协议",
        "multi-embodiment manipulation；failure 分析；跨机器人 benchmark",
        "multi-embodiment IL；failure-aware 训练；benchmark",
        "RoboMIND baseline / multi-embodiment IL",
        "多机器人 action 统一难；版本口径（55k/107k）；digital twin 待验证",
        "失败数据入库 + 多 embodiment 统一协议标杆",
    ],
    # ── 仿真 traj ──
    [
        "—",
        "仿真 task suite / 算法回归专用；sim-real gap 大，不替代真机 Phase 2 建设",
        "LIBERO",
        "UT Austin / Stanford 等",
        "2023 (NeurIPS)",
        "开源",
        "130 tasks / 4 suites；每任务含 human teleop demos",
        "仿真机器人轨迹（人在仿真中 teleop 的 demonstration）",
        "是",
        "仿真第三人称 image/state",
        "仿真",
        "Panda/Franka 类仿真臂",
        "sim RGB/state、action、language task suite",
        "有",
        "否（benchmark 以成功 demo 为主）",
        "sim EEF delta / joint / sim frame (~7维)",
        "LIBERO 套件",
        "spatial/object/goal/100 四类迁移；lifelong learning 评估",
        "仿真 IL benchmark；lifelong/迁移算法回归（非真机采集路线）",
        "lifelong IL / BC baseline",
        "sim-real gap 大；不代表真机数据链路",
        "内部 task suite 与分布偏移评估设计参考",
    ],
    [
        "—",
        "配合 RoboMIND 真机的 sim 对照补充；不能单独作为 action 空间定义起点",
        "RoboMIND（Isaac Sim 轨迹）",
        "X-Humanoid / RoboMIND 团队",
        "2025",
        "随 RoboMIND 发布",
        "与真机集绑定（digital twin 支持）",
        "仿真机器人轨迹（与真机场景绑定的 sim 复现）",
        "是",
        "仿真渲染视角",
        "仿真",
        "与真机 4 embodiments 对应的 sim 模型",
        "sim traj、language、与真机绑定的 scene meta",
        "有",
        "部分",
        "与真机协议对齐的 sim action",
        "Isaac Sim + RoboMIND 协议",
        "sim-real 对照；multi-embodiment benchmark 补充",
        "仿真 IL / sim-real 研究 / benchmark",
        "RoboMIND benchmark",
        "物理可信度需验证；不能替代真机 action 空间定义",
        "真机+仿真绑定的 benchmark 设计参考",
    ],
    # ── 合成 traj ──
    [
        "—",
        "Phase 5b 结构化 sim 增广；须真机 seed 与 predicate 对齐后再用，非首版路线",
        "MimicGen 合成演示集",
        "NVIDIA / Stanford / UT Austin 等",
        "2023 (CoRL)",
        "开源（方法+生成数据）",
        "由少量 seed demo 扩增，规模取决于任务配置",
        "仿真合成机器人轨迹（算法生成，非人类逐条 teleop）",
        "是",
        "仿真 state/render（RoboSuite 类）",
        "仿真",
        "RoboSuite 等仿真机械臂",
        "state/action traj、subtask boundary、success predicate",
        "弱",
        "自动 predicate 过滤失败 rollout",
        "sim state-action / sim privileged 坐标",
        "RoboSuite / HDF5 等",
        "结构化 pick-place/assembly/rearrangement；长尾物体位姿补充",
        "仿真 IL；sim 增广（真机定标准后的补充）",
        "RoboSuite IL / 仿真 BC",
        "任务须可分子任务；依赖 privileged state；sim-real gap",
        "少 seed demo 扩增长尾的可复现参考",
    ],
    # ── 历史 / 不推荐 ──
    [
        "不推荐",
        "2019 前后旧格式（TFRecord 等）；现代 VLA/IL 需重度重处理，仅作历史对比勿作建设起点",
        "RoboNet",
        "Berkeley / Stanford 等",
        "2019 前后（历史）",
        "开源",
        "15M+ frames / 7 robot platforms",
        "真机机器人遥操作轨迹（早期格式）",
        "是",
        "第三人称 RGB 为主",
        "真机",
        "7 种早期机器人平台",
        "RGB、action（格式与 schema 较旧）",
        "弱/不统一",
        "—",
        "各异 / 各异坐标",
        "TFRecord 等旧格式",
        "早期 multi-robot 视觉表示学习；历史对比",
        "视觉预训练/表示学习（现代 VLA 需重度重处理）",
        "现代 VLA 需重处理",
        "格式老旧；不宜作现代 VLA/IL 直接样板",
        "历史参考；非当前建设起点",
    ],
]

# ---------------------------------------------------------------------------
# 4-链路与工具
# ---------------------------------------------------------------------------
PIPELINE_HEADERS = [
    "落地推荐",
    "选择理由",
    "大类",
    "名称",
    "类别",
    "链路环节",
    "输入",
    "输出",
    "模型参与建链路",
    "对标 Phase",
    "何时采用",
    "主要局限",
    "备注",
]

# 排序：A 建链路（Phase 1→2→5b→5c）→ B 消费端（IL → VLA 桌面 → VLA 人形）
PIPELINE_ROWS = [
    # ── A · 建链路 ──
    [
        "★★",
        "Phase 1–2 工程落地首选：episode 录制、存储、训练一条龙，从 0 到 1 最快",
        "A·建链路",
        "LeRobot",
        "数据格式与工具链",
        "episode schema + 录制/训练/export",
        "真机或仿真 raw logs",
        "LeRobot v3 episode 数据集 + 训练 pipeline",
        "无",
        "Phase 1–2",
        "快速自建 dataset、工程落地",
        "与 RLDS/VLA 生态需 converter 桥接",
        "框架非数据集；定义字段而非提供 traj",
    ],
    [
        "★★",
        "Phase 4 多源标准化核心；上 VLA 前须理解 converter + mixture。注意：Figure/Tesla 的闭源 mixture ≠ 本行公开 OXE",
        "A·建链路",
        "Open X-Embodiment (OXE)",
        "聚合与标准化平台",
        "converter + schema + data mixture",
        f"60 个异构源数据集（类型混合，见「{T_DATASET}」各源）",
        "RLDS 统一训练包；RT-1-X/RT-2-X 预训练资产",
        "无（规则 converter；RT 模型为消费端）",
        "Phase 1 + 4",
        "多源/多机器人数据需要统一 schema 与 mixture 权重时",
        "源数据质量不均；action adapter 难；mixture 权重隐性关键",
        f"学术/中小团队互操作层；{T_INDUSTRY}: 巨头多为闭源版 Phase4（勿照搬）",
    ],
    [
        "★",
        "双臂精细 teleop 硬件参考；回答「怎么采」而非训练算法，按需选型（行业锚点：图5 宇树/双臂团队）",
        "A·建链路",
        "ALOHA",
        "采集硬件与 teleop 范式",
        "leader-follower 双臂遥操作",
        "操作者意图 + 双臂 leader 动作",
        "自建双臂 demonstration（规模取决于团队采集）",
        "无",
        "Phase 2",
        "双臂精细操作、低成本 IL 硬件选型",
        "硬件搭建成本；非大规模公开统一数据集",
        f"区别于 ACT：ALOHA 是「怎么采」；{T_INDUSTRY}: 宇树路径常用此类 teleop",
    ],
    [
        "—",
        "Phase 5b 结构化 sim 增广；真机 seed 与 predicate 对齐后再引入",
        "A·建链路",
        "MimicGen（方法）",
        "仿真数据扩增方法",
        "seed demo → subtask 变换 → 合成 traj",
        "少量人类/脚本 seed demonstration + 仿真环境",
        f"合成 demonstration（产物见「{T_DATASET}」MimicGen 合成演示集）",
        "中（几何 subtask stitching；非 LLM 生成）",
        "Phase 5b",
        "真机已有标准与 predicate，需补结构化任务长尾位姿",
        "任务须可分解；依赖 sim privileged state",
        f"方法与产物分行：方法在本表，数据在「{T_DATASET}」",
    ],
    [
        "—",
        "Phase 5c 场景/asset 补全；产出非 traj，sim-real 待验证，非首版必选项",
        "A·建链路",
        "SimFoundry / PolaRiS",
        "Real2Sim 链路",
        "视频 → sim-ready 场景/asset",
        "人类第三人称 RGB 视频（exo，非 robot traj）",
        "mesh / 3DGS / physics meta / digital twin（非 trajectory 数据集）",
        "强（分割/深度/重建/物理稳定等视觉模型 pipeline）",
        "Phase 5c",
        "需补场景/物体分布、少采真机场景时",
        "产出是场景资产不是 demo；sim-real 待验证；不能直接当 IL 训练集",
        "输入=exo 视频；输出≠ego 数据集≠mocap",
    ],
    # ── B · 消费端 ──
    [
        "★★",
        "单任务 IL 默认首选；有 teleop 数据即可训，门槛最低、落地最快（行业工程默认见图5 宇树）",
        "B·消费端",
        "ACT",
        "模仿学习训练算法",
        "—（消费 demo 数据，不产出数据）",
        "RGB + proprio + action chunk 对齐的 episode",
        "action chunking 策略模型",
        "—",
        "—",
        "单任务/双臂 IL；已有 teleop 数据后",
        "开环 chunk 误差；需足够 demo 覆盖",
        f"不进「{T_DATASET}」；与 LeRobot 配合；{T_INDUSTRY}: 宇树路径",
    ],
    [
        "★",
        "Bridge 生态 language-conditioned 桌面操作；action 空间与 BridgeData 高度相关",
        "B·消费端",
        "RT-1 / RT-1-X",
        "VLA/IL 模型（消费端）",
        "—",
        "language + Bridge 系 action 或 RLDS adapter 数据",
        "language-conditioned 操作策略",
        "无",
        "Phase 3–4",
        "桌面 language-conditioned 操作；Bridge 生态",
        "跨 embodiment 需 adapter",
        "与 BridgeData V2 action 空间高度相关",
    ],
    [
        "★",
        "cross-embodiment VLA 框架；多机器人预训练/微调备选，需 RLDS + language",
        "B·消费端",
        "Octo",
        "VLA 模型/训练框架（消费端）",
        "—（定义数据需满足的条件）",
        "RLDS + language + embodiment meta + 多视角 RGB",
        "cross-embodiment VLA 策略",
        "无",
        "Phase 4+",
        "有 converter 与足够 language 数据后做 VLA",
        "fine-tune 仍要自有真机数据",
        "反向约束：数据需 language 与 embodiment 字段",
    ],
    [
        "★★",
        "当前 VLA 微调社区最活跃之一；Bridge/OXE 经 RLDS 后单团队落地首选。Figure 类「VLA 叙事」≠ 本行公开权重/数据自动可得",
        "B·消费端",
        "OpenVLA",
        "VLA 模型/训练框架（消费端）",
        "—",
        "大规模 RLDS mixture + language",
        "通用 language-conditioned VLA",
        "无",
        "Phase 4+",
        "大规模 VLA 预训练/微调",
        "算力与数据规模大",
        f"常用 OXE/Bridge；{T_INDUSTRY}: Figure 叙事对标本行能力层级，数据多为闭源",
    ],
    [
        "—",
        "人形/灵巧手专用 VLA；仅智元类人形硬件团队考虑，非通用默认（行业锚点：图5 智元/Sanctuary）",
        "B·消费端",
        "GO-1 / ViLLA",
        "人形/灵巧手 VLA（消费端）",
        "—",
        "RGBD、tactile、高维 dexterous action",
        "人形/灵巧手策略",
        "无",
        "Phase 2–3（人形团队）",
        "AgiBot 类人形硬件与任务",
        "硬件绑定强",
        f"消费 AgiBot World；详见图5 智元「可复用/可借鉴」",
    ],
]

# ---------------------------------------------------------------------------
# 1-自建路线图
# ---------------------------------------------------------------------------
ROADMAP_HEADERS = [
    "优先级",
    "Phase",
    "选择理由",
    "需建设的数据内容类型",
    "建设内容",
    "交付物",
    "对标参考",
    "周期粗估",
    "阻塞下一阶段",
    "验收 Checklist",
]

ROADMAP_ROWS = [
    [
        "P0",
        "Phase 1",
        "一切采集的前提；不定 schema / action 表示，后续清洗与融合成本指数上升",
        "—（先定协议，再采 traj）",
        "统一 episode schema、task taxonomy、action 表示、metadata",
        "数据协议 v1、字段文档、converter 模板",
        f"{T_PIPE}: OXE / LeRobot",
        "2–4 周",
        "是",
        "□ schema 评审 □ 样例 episode □ 1 个 converter 跑通",
    ],
    [
        "P1",
        "Phase 2",
        "默认落地主线；100–500 条真机 traj 即可出 IL baseline，阻塞所有后续阶段",
        "真机机器人遥操作轨迹",
        "单站 teleop + 回放 + 入库质检",
        "采集/回放工具；首版 100–500 条 traj",
        f"{T_DATASET}: BridgeData V2 / DROID  |  {T_PIPE}: ALOHA  |  {T_INDUSTRY}: 宇树路径",
        "4–8 周",
        "是",
        "□ 可录可回放 □ 成功率可测 □ 100+ 条入库 □ 场景/任务分布记录",
    ],
    [
        "P2",
        "Phase 3",
        "首版数据入库后再做；提升可维护性与版本管理，不阻塞 Phase 2 出 baseline",
        "真机遥操作轨迹（质量增强）",
        "自动质检、质量评分、版本管理、分布监控",
        "quality dashboard、dataset registry",
        f"{T_DATASET}: AgiBot / RH20T / DROID  |  {T_INDUSTRY}: 智元 QA（★★）",
        "4–6 周",
        "否",
        "□ 6 维质量分 □ 版本可回滚 □ failure_type taxonomy",
    ],
    [
        "P3",
        "Phase 4",
        "多源/VLA 前置；算力、converter、自有 fine-tune 集就绪后再上，非首月必做；勿把巨头闭源 mixture 当成公开 OXE",
        "多源 traj（混合类型经 converter 统一）",
        "多源融合、mixture 权重、训练集构建",
        "mixture config、baseline training set、eval split",
        f"{T_PIPE}: OXE / RT-1-X  |  {T_INDUSTRY}: 闭源 Phase4 见特斯拉/Figure（勿照搬）",
        "4–8 周",
        "否",
        "□ converter 齐全 □ mixture 文档 □ hold-out eval",
    ],
    [
        "P3",
        "Phase 5a",
        "failure 专项扩展；IL baseline 稳定、success/failure 同 schema 后再建",
        "真机 traj + failure 样本",
        "失败入库 + targeted 补采",
        "failure dataset、补采队列",
        f"{T_DATASET}: RoboMIND（failure）",
        "持续",
        "否",
        "□ failure_type □ 与成功同 schema □ 补采优先级",
    ],
    [
        "P3",
        "Phase 5b",
        "结构化 sim 增广；真机 predicate 与 seed demo 对齐后补充长尾位姿",
        "仿真合成机器人轨迹",
        "MimicGen 式结构化 sim 扩增",
        "synthetic traj、predicate 报告",
        f"{T_DATASET}: MimicGen 合成集  |  {T_PIPE}: MimicGen 方法",
        "4–6 周",
        "否",
        "□ seed≤20 可扩增 □ predicate 与真机一致 □ sim-real 抽检",
    ],
    [
        "P3",
        "Phase 5c",
        "场景/asset 补全；少采真机场景时的 long-term 选项，产出非 traj",
        "场景/asset（非 traj）",
        "Real2Sim 场景生成",
        "sim-ready 场景库、digital twin meta",
        f"{T_PIPE}: SimFoundry",
        "6–12 周",
        "否",
        "□ 视频→场景 MVP □ 物理稳定 □ scene_id 与真机绑定",
    ],
    [
        "—",
        "训练选型",
        "见「2-训练范式速查」；默认 ACT 出 baseline → 有 language 需求再 VLA 微调",
        "依数据内容类型匹配",
        "选定 IL / VLA / Benchmark 路线",
        "baseline 报告、eval 协议",
        f"{T_PARADIGM}  |  {T_INDUSTRY}: 先看 ★★/★ 与勿照搬",
        "随 Phase 2+",
        "否",
        "□ 消费端明确 □ train/eval 场景分离 □ 指标可复现 □ 已查阅图5可复用/可借鉴/勿照搬",
    ],
]

# ---------------------------------------------------------------------------
# 2-训练范式速查
# ---------------------------------------------------------------------------
PARADIGM_HEADERS = [
    "推荐阶段",
    "落地推荐",
    "选择理由",
    "训练范式",
    "需要什么数据内容类型",
    "数据门槛摘要",
    "推荐公开数据",
    "常用消费模型/框架",
    "典型任务",
    "主要限制",
]

# 排序：数据门槛从低到高；推荐阶段与「1-自建路线图」Phase / 优先级闭环
PARADIGM_ROWS = [
    [
        "Phase 2 · P1",
        "★★",
        "门槛最低、落地最快；有 teleop 数据即可训，默认首条训练路线（行业工程默认见图5 宇树路径）",
        "单任务 IL（如 ACT）",
        "真机机器人遥操作轨迹",
        "RGB + proprio + action 对齐；单 embodiment；可选 task 文本",
        f"自建为主；硬件参考 {T_PIPE}: ALOHA",
        f"{T_PIPE}: ACT / Diffusion Policy",
        "单站精细操作、双臂操作",
        f"chunk 开环误差；场景覆盖需足够 demo；行业对标 {T_INDUSTRY}: 宇树",
    ],
    [
        "Phase 2–3 · P1",
        "★★",
        "有 language 任务时的现实 VLA 路径；公开数据 + 自有 fine-tune，不必从零预训练（Figure 类叙事≠公开 OXE）",
        "VLA / 策略微调（单团队落地）",
        "真机遥操作轨迹",
        "与预训练 action 空间一致或可控 adapter + 自有 fine-tune 集",
        f"{T_DATASET}: BridgeData V2 + 自建真机 traj",
        f"{T_PIPE}: Octo / OpenVLA / RT-1",
        "桌面/固定站房任务落地",
        f"公开数据不能替代自有 fine-tune；行业对标 {T_INDUSTRY}: Figure/OpenVLA 类",
    ],
    [
        "可选 · —",
        "—",
        "算法/迁移评估专用；不采真机也能跑，但不替代 Phase 2 真机建设",
        "仿真 Benchmark / 算法回归",
        "仿真机器人轨迹（人采 demo）",
        "固定 task suite、可复现 sim 环境",
        f"{T_DATASET}: LIBERO",
        "lifelong IL baseline",
        "迁移/终身学习评估",
        "sim-real gap；不替代真机建设",
    ],
    [
        "Phase 4 · P3",
        "★",
        "大算力 + 多源 RLDS；单团队通常 fine-tune 即可，全量预训练非默认（巨头闭源 mixture 见图5勿照搬）",
        "VLA 预训练 / 大规模微调",
        "真机遥操作轨迹（多源经 converter）",
        "RLDS、language、embodiment meta、多视角 RGB",
        f"{T_PIPE} OXE 源之一 → {T_DATASET}: BridgeData V2 / DROID",
        f"{T_PIPE}: Octo / OpenVLA / RT-1-X",
        "language-conditioned 跨场景操作",
        f"算力大；action adapter；数据质量敏感；{T_INDUSTRY}: 特斯拉/Figure≠公开OXE",
    ],
    [
        "Phase 5b · P3",
        "—",
        "真机标准与 predicate 定好后补长尾；不能替代 action 空间定义",
        "Sim 增广 IL",
        "仿真合成机器人轨迹",
        "少量 seed + 可分解 subtask + success predicate",
        f"{T_DATASET}: MimicGen 合成演示集",
        "RoboSuite IL",
        "结构化 pick-place/assembly",
        "不能替代真机 action 空间定义",
    ],
    [
        "Phase 5a · P3",
        "—",
        "跨机器人 + failure 分析扩展；统一协议成本高，baseline 后考虑",
        "Multi-embodiment / failure-aware",
        "真机 traj + failure；可选 sim 绑定 traj",
        "统一 teleop 协议 + failure taxonomy",
        f"{T_DATASET}: RoboMIND（真机+failure / Sim）",
        "RoboMIND baseline",
        "跨机器人 benchmark",
        "action 统一成本高",
    ],
    [
        "Phase 3+ · P2",
        "—",
        "力/触觉/多传感器专项；工程同步门槛高，接触丰富任务才值得投入（灵巧手路径见图5 Sanctuary）",
        "多模态 / 接触丰富 IL",
        "真机遥操作轨迹 + FT/tactile/audio",
        "多传感器时间同步与标定",
        f"{T_DATASET}: RH20T",
        "多模态 IL 研究框架",
        "接触/力控/插孔等",
        f"工程复杂；非 VLA 开箱即用；{T_INDUSTRY}: Sanctuary/GO-1 行",
    ],
    [
        "Phase 5c · P3",
        "—",
        "补场景/物体分布；产出是 asset 不是 traj，不直接用于 IL 训练",
        "Real2Sim 场景补全",
        "人类 exo 视频（输入）→ 场景 asset（输出）",
        "视频重建 + 物理稳定；非 trajectory",
        f"—（产出链路见 {T_PIPE}: SimFoundry）",
        "接 MimicGen/scripted/RL 产 traj",
        "补场景/物体分布",
        "不直接用于 IL 训练",
    ],
]


# ---------------------------------------------------------------------------
# 5-行业对标（人形）
# 公司名「中文 英文」；保留在建链路+技术路线；重点列用色/字重强调
# ---------------------------------------------------------------------------
INDUSTRY_BANNER = (
    "读表：①落地借鉴 → ②在建链路/技术路线（行业在做什么）→ ③可复用/可借鉴/勿照搬 → ④建议动作。"
    "★★主对标｜★学一环｜—知存在。默认宇树路径；公开样板智元；勿把特斯拉/Figure 闭源写成上 OXE。"
    "口径含公开与推断，以公开度为准；截至 2026-Q2。"
)

INDUSTRY_BANNER2 = (
    "路线簇｜①公开工厂→人形VLA：智元｜②硬件生态→自建IL：宇树=工程默认｜"
    "③交付型teleop：银河通用/傅利叶/优必选｜④海外闭源VLA：Figure等｜⑤车企垂直：特斯拉/小鹏"
)

INDUSTRY_HEADERS = [
    "落地借鉴",
    "公司/产品",
    "大类",
    "选择理由",
    "在建链路",
    "技术路线",
    "主 Phase",
    "消费端对标",
    "可复用",
    "可借鉴",
    "勿照搬",
    "建议动作",
    "公开度",
]

# 排序：中国·可对标 → 中国·知存在 → 海外·头部 → 海外/车企
INDUSTRY_ROWS = [
    # ── 中国·可对标 ──
    [
        "★★",
        "智元 AgiBot",
        "中国·可对标",
        "国内最完整「采–标–训」公开样板",
        "真机 teleop 数据工厂 → 结构化标注 → 人形 VLA（GO-1 / World）",
        "Teleop工厂; 人形VLA; 部分开源",
        "Phase 2–3",
        f"{T_PIPE}: GO-1",
        "AgiBot World 子集; taxonomy/QA 字段; GO-1 公开材料",
        "数据工厂组织; 多轮QA; skill/keyframe 切分",
        "百万规模本身; 与G1强绑定的 action 直混",
        "Phase2 后：Phase3 抄其 QA/skill；桌面勿改默认 ACT",
        "部分开源",
    ],
    [
        "★",
        "宇树 Unitree",
        "中国·可对标",
        "硬件生态+社区 teleop+ACT，工程最短路径",
        "本体+开发生态自建 Phase2；消费端常见 ACT/Diffusion/自研 IL",
        "硬件生态; Teleop自建; ACT/IL",
        "Phase 2",
        f"{T_PIPE}: ACT / Diffusion",
        "本体/SDK/仿真; 社区 LeRobot/ACT 教程",
        "先硬件后数据; LeRobot schema 降融合成本",
        "有G1≠已有智元级数据工厂; 等官方开源替代自建",
        "默认：表1 Phase2 + 表4 LeRobot/ACT（宇树路径）",
        "硬件公开; 数据自建",
    ],
    [
        "★",
        "银河通用 Galbot",
        "中国·可对标",
        "具身操作+交付叙事清晰，偏自研数据闭环",
        "真机 teleop + 场景交付 → 自研操作策略/具身模型",
        "交付型Teleop; 自研消费端",
        "Phase 2–3",
        "自研 IL / VLA",
        "无稳定公开 traj",
        "场景交付下的采集 checklist; 数据权属条款",
        "演示规模与未公开「大模型」口径",
        "商务对标可谈场景；技术主线仍用智元/宇树",
        "官网/展会; 数据闭源",
    ],
    [
        "★",
        "傅利叶 Fourier",
        "中国·可对标",
        "真机交付成熟，公开大数据少",
        "真机交付与行业方案；项目制采集站 + 自研/合作消费端",
        "交付型Teleop; 真机方案",
        "Phase 2–3",
        "自研 IL / 合作模型",
        "无",
        "项目制采集站节奏; 交付场景任务定义",
        "把交付演示当可复现训练标准",
        "保持表1 自建主线；仅场景参考",
        "产品公开; 数据闭源",
    ],
    [
        "★",
        "优必选 UBTECH",
        "中国·可对标",
        "Walker 场景数据偏项目制，可学场景定义",
        "Walker 场景真机数据 + 项目交付链路 → 场景策略",
        "场景方案; 项目制采集",
        "Phase 2–3",
        "自研 / 合作 VLA·IL",
        "无通用公开操作集",
        "场景 KPI 驱动任务切分; 集成商数据权属",
        "把 Walker 演示当通用家用预训练",
        "非默认路径；业务匹配时再单独立项",
        "案例为主",
    ],
    [
        "★",
        "星尘智能 Astribot",
        "中国·可对标",
        "双臂精细操作叙事强，数据少公开",
        "双臂精细 teleop 采集 → 自研 IL/操作策略",
        "双臂Teleop; 精细操作",
        "Phase 2",
        "自研 IL",
        "无",
        "双臂 teleop 工位与同步要求",
        "未公开模型栈当默认选型",
        "双臂任务时对照表4 ALOHA；数据仍自建",
        "演示/PR; 数据闭源",
    ],
    # ── 中国·知存在 ──
    [
        "—",
        "众擎 / 开普勒 EngineAI / 逐际动力等",
        "中国·知存在",
        "国内玩家多，公开数据工厂少",
        "早期真机演示与量产叙事；公开采集链路未成型",
        "真机演示; 早期交付",
        "Phase 2",
        "自研或未披露",
        "无",
        "跟踪融资与量产节点即可",
        "名单膨胀代替建设决策",
        "不进主对标；有公开集再升格★",
        "碎片信息",
    ],
    [
        "—",
        "小鹏 Iron",
        "车企·闭源",
        "车企具身垂直整合，资产不可用",
        "车企数据体系 + 人形 teleop + 内部模型栈",
        "闭源Mixture; 垂直整合",
        "Phase 2+",
        "端到端自研（闭源）",
        "无",
        "感知与操作数据统一版本治理（思想）",
        "写成「对齐小鹏上 OXE」",
        "只校准天花板；勿替代自建 Phase2",
        "强推断",
    ],
    # ── 海外·头部 ──
    [
        "★",
        "Figure",
        "海外·头部",
        "真机工厂→端到端 VLA 天花板叙事",
        "量产向真机 teleop + 产线任务（BotQ 等）→ Helix 类 VLA 消费",
        "Teleop工厂; VLA; 闭源",
        "Phase 2–3",
        "能力近 OpenVLA；权重不公开",
        "几乎无",
        "采集站吞吐; 策略与产线 KPI 绑定",
        "闭源 mixture 规模; PR 当可复现标准",
        "学工厂化组织；VLA 走表2 公开微调+自建数据",
        "叙事/演示",
    ],
    [
        "★",
        "1X",
        "海外·头部",
        "人在环 teleop 飞轮清晰",
        "人在环遥操作/协助操作产数据 → 再训家用操作策略",
        "Teleop飞轮; 逐步自主",
        "Phase 2",
        "自研 IL → 轻量 VLA",
        "无",
        "人在环补采队列; 协助→自主配比",
        "家用合规细节与具体模型栈",
        "Phase2 嵌入「人在环补采」；勿等全自主再采",
        "产品/PR; 数据闭源",
    ],
    [
        "★",
        "Sanctuary AI",
        "海外·头部",
        "高维灵巧手 teleop 专项",
        "高带宽遥操作采精细操作 → 再自主化",
        "灵巧手Teleop; 高维action",
        "Phase 2",
        "近 GO-1/多模态接触 IL",
        "无官方集；方法可对 RH20T",
        "高维手标定同步; 接触示教工时",
        "整机液压手当通用人形默认",
        "仅接触任务：先表2 多模态+RH20T",
        "演示/论文; 数据闭源",
    ],
    # ── 海外/车企·知存在 ──
    [
        "—",
        "Agility Robotics",
        "海外·场景",
        "仓储垂直，通用人形参考有限",
        "仓储场景真机日志 + 遥操作/辅助 → 场景策略",
        "场景垂直; 真机日志",
        "Phase 2–3",
        "自研 IL/RL 混合",
        "无",
        "任务定义先于模型; 现场采集合规",
        "仓储 Digit 数据当家用预训练",
        "仅业务=仓储物流时立项",
        "产品案例",
    ],
    [
        "—",
        "特斯拉 Optimus / 车企具身",
        "车企·闭源",
        "闭源 Phase4 mixture 天花板，资产不可用",
        "车规/工厂数据体系 + 真机 teleop + 内部仿真/孪生",
        "闭源Mixture; 仿真辅路",
        "Phase 2 / 内部4",
        "端到端自研（非公开 OpenVLA）",
        "无（严禁表3 伪造成开源）",
        "垂直整合时版本治理; 仿真作辅",
        "车队数据量; 「对齐 Tesla 上 OXE」",
        "Phase4 仅算力+自有 fine-tune 就绪后用公开 OXE",
        "强推断",
    ],
]


def _tab_color(ws, rgb: str) -> None:
    ws.sheet_properties.tabColor = rgb


def main() -> None:
    wb = Workbook()

    # Tab1 — 自建路线图
    ws_road = wb.active
    ws_road.title = T_ROAD
    style_table(
        ws_road,
        ROADMAP_HEADERS,
        ROADMAP_ROWS,
        col_widths=COL_WIDTHS[T_ROAD],
        freeze="D2",
        priority_col=1,
    )
    _tab_color(ws_road, C_ACCENT)

    # Tab2 — 训练范式
    ws_par = wb.create_sheet(T_PARADIGM)
    style_table(
        ws_par,
        PARADIGM_HEADERS,
        PARADIGM_ROWS,
        col_widths=COL_WIDTHS[T_PARADIGM],
        freeze="D2",
        name_col=4,
        recommend_col=2,
    )
    _tab_color(ws_par, "6947B5")

    # Tab3 — 数据集目录
    ws_data = wb.create_sheet(T_DATASET)
    style_table(
        ws_data,
        DATASET_HEADERS,
        DATASET_ROWS,
        col_widths=COL_WIDTHS[T_DATASET],
        freeze="D2",
        name_col=3,
        recommend_col=1,
        section_starts=[6, 8, 9],
    )
    _tab_color(ws_data, "138A5B")

    # Tab4 — 链路与工具
    ws_pipe = wb.create_sheet(T_PIPE)
    style_table(
        ws_pipe,
        PIPELINE_HEADERS,
        PIPELINE_ROWS,
        col_widths=COL_WIDTHS[T_PIPE],
        freeze="D2",
        name_col=4,
        recommend_col=1,
        section_starts=[6],
    )
    _tab_color(ws_pipe, "C97815")

    # Tab5 — 行业对标（人形）
    ws_ind = wb.create_sheet(T_INDUSTRY)
    style_table(
        ws_ind,
        INDUSTRY_HEADERS,
        INDUSTRY_ROWS,
        col_widths=COL_WIDTHS[T_INDUSTRY],
        freeze="C4",
        name_col=2,
        recommend_col=1,
        category_col=3,
        section_starts=[1, 7, 9, 12],
        banner=INDUSTRY_BANNER,
        banner2=INDUSTRY_BANNER2,
        banner_height=42,
        banner2_height=36,
        data_row_height=46,
        auto_filter=True,
        header_fills=INDUSTRY_HEADER_FILLS,
        emphasis_cols=INDUSTRY_EMPHASIS,
    )
    _tab_color(ws_ind, "8B3A3A")

    wb.save(OUTPUT)
    print(f"已生成: {OUTPUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
